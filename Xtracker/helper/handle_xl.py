
from openpyxl import Workbook



class XL_Man:
    
    
    def __init__(self):
        self.wb = Workbook()
        self.sheets =[]
        
    def add_sheet(self,sheet_name):
        setattr(self, "sheet"+str(len(self.sheets)+1), self.wb.create_sheet(sheet_name))
        self.sheets.append(sheet_name)
        
    def _get_sheet(self,sheet_name):
        if sheet_name in self.sheets:
            sheet = getattr(self, "sheet"+str(self.sheets.index(sheet_name)+1))
            return sheet
        else:
            return None
        
    def write(self,sheet_name,row,col,data):
        sheet = self._get_sheet(sheet_name)
        if sheet:
            sheet.cell(row=row,column=col,value=data)
        else:
            print("Invalid sheet")
            
    def save(self,xls_name):
        self.wb.save(xls_name)
        